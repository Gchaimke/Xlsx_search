import os
import datetime

from openpyxl import Workbook
from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox
from tkinter.ttk import *
from tkinter import filedialog

working_folder = os.getcwd()
window = Tk()
input_txt = Entry(window)
input_txt.insert(0, "mio")
progress = Progressbar(window, orient=HORIZONTAL,
                       length=250)
foldersList = Listbox(window, width=80)

def search_files():
    my_files_list = []
    for folder_path in foldersList.get(0, END):
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if file.endswith(".xlsx"):
                    if os.path.basename(file)[0] != "~":
                        my_files_list.append(os.path.join(root, file))
    return my_files_list


def is_int(s):
    try:
        int(s)
        return True
    except ValueError:
        return False


def save_workbook(rows):
    workbook = Workbook()
    sheet = workbook.active
    now = datetime.datetime.now()
    time_stamp = now.strftime("%d_%m_%y__%H_%M_%S")
    file_name = "search_"+time_stamp+".xlsx"

    for row in rows:
        if len(row) > 0:
            sheet.append(row)

    sheet.freeze_panes = 'B2'
    # print(f"output rows {sheet.max_row}")
    file_path = os.path.join(working_folder, file_name)
    print(file_path)
    progress['value'] = 0
    workbook.save(filename=file_path)
    os.startfile(working_folder)


def open_workbook(file_path):
    try:
        folder, file_name = os.path.split(file_path)
        file_name = os.path.splitext(file_name)[0]
        workbook = load_workbook(file_path, read_only=True)
        sheet = workbook.active
        # print(f"input rows {sheet.max_row}")
    except PermissionError as e:
        print(f"{e}")
        messagebox.showwarning(
            'Error', f"{e} \n File opened in other app?", icon='error')
    except Exception as e:
        print(f"{e}")
        messagebox.showwarning(
            'Error', f"{e} \n File is broken?", icon='error')
        return

    rows = []
    file_rows = sheet.iter_rows(max_col=4, values_only=True)
    search = input_txt.get().lower()
    for row in file_rows:
        try:
            if("pn" in f"{row[0]}".lower()):
                continue
            for cell in row:
                if(search in f"{cell}".lower()):
                    row = (file_name,)+row
                    rows.append(row)
        except Exception as e:
            print(f"Error read row: {e} {row}")
            continue
    return rows


def run_app():
    global working_folder
    search = input_txt.get()
    if(search==""):
        messagebox.showwarning('Error', f"Search is empty", icon='error')
        return
    file_list = search_files()
    print(f"Total files:{len(file_list)}")
    file_rows = [("SYSTEM", "PN", "SN")]
    try:
        progress_step = 100/len(file_list)
        for file in file_list:
            if(progress['value'] > 100):
                progress['value'] = 0
            progress['value'] += progress_step
            window.update_idletasks()
            if("tmp" in f"{file}".lower()):
                continue
            file_rows += open_workbook(file)
        save_workbook(file_rows)
        return
    except ZeroDivisionError as e:
        print(f"{e}")
        messagebox.showwarning('Error', f"No folders in list", icon='error')
    except Exception as e:
        print(f"{e}")
        messagebox.showwarning('Error', f"{e}", icon='error')


def add_folder():
    foldersList.insert(END, filedialog.askdirectory(title="select folder"))
    return


def clear_folder_list():
    foldersList.delete(0, END)
    return


def gui():
    print("Gui start")
    window.title("Search in Excels")
    if os.path.isfile('icon.ico'):
        window.iconbitmap('icon.ico')
    input_lbl = Label(window, text="Search for PN contain: ")
    lbl = Label(window, text="Progress")
    btnStart = Button(window, text='Start', command=run_app)
    btnFolders = Button(window, text='Add Folder', command=add_folder)
    btnClearFolders = Button(window, text='clear', command=clear_folder_list)
    foldersList.grid(row=0, column=0, padx=10, columnspan=2, rowspan=2)
    btnFolders.grid(row=0, column=2, padx=[5, 10])
    btnClearFolders.grid(row=1, column=2, padx=[5, 10])
    input_lbl.grid(row=2, column=0)
    input_txt.grid(row=2, column=1, pady=10, padx=10, columnspan=2, ipadx=100)
    progress.grid(row=3, column=1,  pady=5, padx=5, columnspan=2, ipadx=40)
    lbl.grid(row=3, column=0,  padx=10)
    btnStart.grid(row=4, column=0,  padx=10, pady=10, columnspan=4)
    window.mainloop()


gui()
